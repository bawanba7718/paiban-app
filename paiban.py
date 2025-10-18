import pandas as pd
import datetime
from datetime import datetime, time, timedelta, timezone
import streamlit as st
import openpyxl
from openpyxl import load_workbook
import os
import time as t
import tempfile
from webdav3.client import Client
import threading
import html
import calendar
from datetime import date

# 定义东八区时区（UTC+8）
TZ_UTC_8 = timezone(timedelta(hours=8))

class AgentViewer:
    def __init__(self):
        # 颜色-职位对应关系
        self.color_roles = {
            'FFC000': 'C席',
            'FFEE79': 'C席',
            'E2EFDA': 'C席',
            '91AADF': 'C席',
            'D9E1F2': 'C席',
            'EF949F': 'B席',
            'FADADE': 'B席',
            '8CDDFA': '休',
            'FFFF00': '休',
            'FFFFFF': 'A席',
            'FEE796': 'C席',
        }
        
        # 席位颜色映射
        self.seat_colors = {
            'C席': '#FFC000',
            'B席': '#EF949F',
            'A席': '#FFFFFF'
        }
        
        # 状态图标
        self.status_icons = {
            '搬砖中': '🛠️',
            '干饭中': '🍚',
            '已回家': '🏠',
            '正在路上': '🚗',
            '未排班': '❓',
            '未知班次': '❓'
        }
        
        # 班次时间定义
        self.shift_times = {
            'T1': {'start': time(8, 0), 'end': time(20, 0), 'name': '白班', 
                  'break_start': time(13, 0), 'break_end': time(14, 0)},
            'T2': {'start': time(20, 0), 'end': time(8, 0), 'name': '夜班',
                  'break_start': None, 'break_end': None},
            'M2': {'start': time(8, 0), 'end': time(17, 0), 'name': '早班',
                  'break_start': time(14, 0), 'break_end': time(15, 0)},
            'E2': {'start': time(13, 0), 'end': time(22, 0), 'name': '晚班',
                  'break_start': time(15, 0), 'break_end': time(16, 0)},
            'E3': {'start': time(13, 0), 'end': time(23, 0), 'name': '晚班',
                  'break_start': time(15, 0), 'break_end': time(17, 0)},
            'M1': {'start': time(7, 0), 'end': time(16, 0), 'name': '早班',
                  'break_start': time(12, 0), 'break_end': time(13, 0)},
            'D1': {'start': time(9, 0), 'end': time(18, 0), 'name': '白班',
                  'break_start': time(12, 0), 'break_end': time(13, 0)},
            'D2': {'start': time(10, 0), 'end': time(19, 0), 'name': '白班',
                  'break_start': time(13, 0), 'break_end': time(14, 0)},
            'D3': {'start': time(11, 0), 'end': time(20, 0), 'name': '白班',
                  'break_start': time(15, 0), 'break_end': time(16, 0)},
            'E1': {'start': time(12, 0), 'end': time(21, 0), 'name': '晚班',
                  'break_start': time(15, 0), 'break_end': time(16, 0)},
            'F1': {'start': time(7, 0), 'end': time(13, 0), 'name': '短班',
                  'break_start': None, 'break_end': None},
            'F2': {'start': time(9, 0), 'end': time(16, 0), 'name': '短班',
                  'break_start': time(13, 0), 'break_end': time(14, 0)},
            'F3': {'start': time(17, 0), 'end': time(23, 0), 'name': '短班',
                  'break_start': None, 'break_end': None},
            'H1': {'start': time(7, 0), 'end': time(11, 0), 'name': '半日班',
                  'break_start': None, 'break_end': None},
            'H2': {'start': time(16, 0), 'end': time(20, 0), 'name': '半日班',
                  'break_start': None, 'break_end': None},
        }

    def get_work_status(self, shift_code, seat, color_code, check_time=None):
        """获取工作状态"""
        if not shift_code or str(shift_code).strip() == '':
            return "未排班", "#BFBFBF", seat
            
        shift_code = str(shift_code).strip()
        main_shift = None
        
        # 提取主班次
        for s in sorted(self.shift_times.keys(), key=lambda x: len(x), reverse=True):
            if s in shift_code:
                main_shift = s
                break
                
        if not main_shift:
            return "未知班次", "#BFBFBF", seat
            
        # 复制基础班次时间
        shift = self.shift_times[main_shift].copy()
        
        # A席M2统一休13:00-14:00
        if seat == 'A席' and main_shift == 'M2':
            shift['break_start'] = time(13, 0)
            shift['break_end'] = time(14, 0)
        
        # FEE796颜色T1班次特殊规则
        elif color_code == 'FEE796' and main_shift == 'T1':
            shift['break_start'] = time(14, 0)
            shift['break_end'] = time(15, 0)
            
            check_time = check_time or datetime.now(TZ_UTC_8).time()
            if time(17, 0) <= check_time < time(20, 0):
                seat = 'A席'
            else:
                seat = 'C席'
        
        # 其他特殊规则
        elif seat == 'B席' and color_code == 'EF949F' and main_shift == 'T1':
            shift['break_start'] = time(14, 0)
            shift['break_end'] = time(15, 0)
        
        elif seat == 'A席' and main_shift == 'D2':
            shift['break_start'] = time(14, 0)
            shift['break_end'] = time(15, 0)
        
        elif seat == 'C席' and color_code == 'FFC000' and main_shift == 'T1':
            shift['break_start'] = time(13, 0)
            shift['break_end'] = time(14, 0)
            
        elif seat == 'C席' and color_code == 'D9E1F2':
            shift['break_start'] = time(14, 0)
            shift['break_end'] = time(15, 0)
        
        elif seat == 'C席' and color_code == 'E2EFDA' and main_shift == 'T1':
            shift['break_start'] = time(14, 0)
            shift['break_end'] = time(15, 0)
        
        elif seat == 'C席' and color_code == 'E2EFDA' and main_shift == 'M2':
            shift['break_start'] = time(14, 0)
            shift['break_end'] = time(15, 0)
        
        elif seat == 'B席' and color_code == 'FADADE' and main_shift == 'M2':
            shift['break_start'] = time(13, 0)
            shift['break_end'] = time(14, 0)
            
        elif seat == 'A席' and main_shift == 'T1':
            shift['break_start'] = time(14, 0)
            shift['break_end'] = time(15, 0)
            
        # 使用东八区时间
        check_time = check_time or datetime.now(TZ_UTC_8).time()
        
        # 解构时间参数
        start, end = shift['start'], shift['end']
        break_start, break_end = shift.get('break_start'), shift.get('break_end')
        
        # 判断是否在工作时间内
        is_night_shift = main_shift == 'T2'
        in_work_time = False
        
        if is_night_shift:
            # 夜班：20:00-次日08:00
            in_work_time = (check_time >= start) or (check_time < end)
        else:
            # 白班/早班：正常时间范围
            in_work_time = start <= check_time < end
            
        # 判断是否在上班路上
        is_on_the_way = False
        if not in_work_time and not is_night_shift:
            is_on_the_way = check_time < start
            
        # 修正：当天的T2班次在非工作时间显示为"正在路上"
        if is_night_shift and not in_work_time:
            is_on_the_way = True
            
        # 休息时间判断
        in_break_time = False
        if break_start and break_end and in_work_time:
            if break_start < break_end:
                in_break_time = break_start <= check_time < break_end
            else:
                in_break_time = check_time >= break_start or check_time < break_end
        
        # 确定最终状态
        if is_on_the_way:
            return "正在路上", "#BFBFBF", seat
        elif not in_work_time:
            return "已回家", "#BFBFBF", seat
        elif in_break_time:
            return "干饭中", "orange", seat
        else:
            return "搬砖中", "green", seat

    def get_cell_color(self, cell):
        try:
            if cell and cell.fill and cell.fill.start_color:
                color = cell.fill.start_color.rgb
                if color:
                    color_str = str(color).upper()
                    if color_str.startswith('FF'):
                        color_str = color_str[2:]
                    elif len(color_str) == 8:
                        color_str = color_str[2:]
                    return color_str if len(color_str) == 6 else "FFFFFF"
            return "FFFFFF"
        except:
            return "FFFFFF"
    
    def load_schedule_with_colors(self, file_path, target_date):
        try:
            if not os.path.exists(file_path):
                st.error(f"文件不存在: {file_path}")
                return None
                
            wb = load_workbook(file_path, data_only=True)
            if '全部排班' not in wb.sheetnames:
                st.error("工作表 '全部排班' 不存在")
                return None
            
            main_sheet = wb['全部排班']
            df_main = pd.read_excel(file_path, sheet_name='全部排班')
            
            target_date_str = target_date.strftime('%Y-%m-%d')
            today_col_idx = None
            
            for idx, col in enumerate(df_main.columns):
                col_str = str(col)
                if (target_date_str in col_str or 
                    target_date.strftime('%m-%d') in col_str or
                    target_date.strftime('%Y/%m/%d') in col_str or
                    target_date.strftime('%m/%d') in col_str):
                    today_col_idx = idx
                    break
            
            if today_col_idx is None:
                st.warning(f"未找到 {target_date_str} 的排班列，可能该日期无排班数据")
                return pd.DataFrame()
            
            color_data = []
            for row_idx, row in enumerate(main_sheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    if len(row) < 4:
                        continue
                        
                    name = str(row[3].value).strip() if row[3].value else ''
                    if not name:
                        continue
                    
                    shift_code = ""
                    color_code = "FFFFFF"
                    if len(row) > today_col_idx:
                        shift_cell = row[today_col_idx]
                        shift_code = str(shift_cell.value).strip() if shift_cell.value else ""
                        color_code = self.get_cell_color(shift_cell)
                    
                    if (not shift_code or 
                        shift_code.strip() in ['', '休', '休息']):
                        continue
                    
                    seat = self.color_roles.get(color_code, 'A席')
                    
                    person_info = {
                        'name': name,
                        'id': str(row[2].value).strip() if row[2].value else '',
                        'workplace': str(row[0].value).strip() if row[0].value else '',
                        'shift': shift_code,
                        'color': color_code,
                        'seat': seat,
                        'status': '',
                        'status_color': '',
                        'actual_seat': seat,
                        'date': target_date
                    }
                    
                    color_data.append(person_info)
                except Exception as e:
                    continue
            
            return pd.DataFrame(color_data)
            
        except Exception as e:
            st.error(f"加载数据失败: {str(e)}")
            return None
    
    def load_agent_month_schedule(self, file_path, agent_id, year, month):
        """加载指定人员当月的排班数据"""
        try:
            if not os.path.exists(file_path):
                st.error(f"文件不存在: {file_path}")
                return None
                
            wb = load_workbook(file_path, data_only=True)
            if '全部排班' not in wb.sheetnames:
                st.error("工作表 '全部排班' 不存在")
                return None
            
            main_sheet = wb['全部排班']
            df_main = pd.read_excel(file_path, sheet_name='全部排班')
            
            # 查找指定人员
            agent_row = None
            for row_idx, row in enumerate(main_sheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    if len(row) < 4:
                        continue
                        
                    current_id = str(row[2].value).strip() if row[2].value else ''
                    if current_id == agent_id:
                        agent_row = row
                        break
                except:
                    continue
            
            if not agent_row:
                st.error(f"未找到工号为 {agent_id} 的人员")
                return None
            
            # 获取人员基本信息
            agent_info = {
                'name': str(agent_row[3].value).strip() if agent_row[3].value else '',
                'id': agent_id,
                'workplace': str(agent_row[0].value).strip() if agent_row[0].value else '',
            }
            
            # 获取当月所有日期的排班
            month_schedule = []
            
            # 遍历所有列，查找日期列
            for col_idx, col in enumerate(df_main.columns):
                col_str = str(col)
                try:
                    # 尝试解析日期
                    col_date = None
                    if '-' in col_str:
                        parts = col_str.split('-')
                        if len(parts) >= 2:
                            month_part = parts[-2] if len(parts) > 2 else parts[0]
                            day_part = parts[-1]
                            if month_part.isdigit() and day_part.isdigit():
                                col_date = date(year, int(month_part), int(day_part))
                    elif '/' in col_str:
                        parts = col_str.split('/')
                        if len(parts) >= 2:
                            month_part = parts[-2] if len(parts) > 2 else parts[0]
                            day_part = parts[-1]
                            if month_part.isdigit() and day_part.isdigit():
                                col_date = date(year, int(month_part), int(day_part))
                    
                    # 如果是目标月份的日期
                    if col_date and col_date.year == year and col_date.month == month:
                        if col_idx < len(agent_row):
                            shift_cell = agent_row[col_idx]
                            shift_code = str(shift_cell.value).strip() if shift_cell.value else ""
                            color_code = self.get_cell_color(shift_cell)
                            
                            if shift_code and shift_code.strip() not in ['', '休', '休息']:
                                seat = self.color_roles.get(color_code, 'A席')
                                
                                day_schedule = {
                                    'date': col_date,
                                    'shift': shift_code,
                                    'color': color_code,
                                    'seat': seat,
                                    'weekday': col_date.strftime('%A')
                                }
                                month_schedule.append(day_schedule)
                except:
                    continue
            
            return {
                'agent_info': agent_info,
                'month_schedule': sorted(month_schedule, key=lambda x: x['date'])
            }
            
        except Exception as e:
            st.error(f"加载月度排班失败: {str(e)}")
            return None
    
    def categorize_by_seat(self, df, check_time=None):
        result = {'A席': [], 'B席': [], 'C席': []}
        if df is None or df.empty:
            return result
        
        for _, person in df.iterrows():
            status, status_color, actual_seat = self.get_work_status(
                person['shift'], 
                person['seat'], 
                person['color'],
                check_time
            )
            person['status'] = status
            person['status_color'] = status_color
            person['actual_seat'] = actual_seat
            
            seat = actual_seat
            if seat in result:
                result[seat].append(person)
            else:
                result['A席'].append(person)
        
        # 排序逻辑
        status_priority = {
            '搬砖中': 3,
            '干饭中': 2,
            '正在路上': 1,
            '已回家': 0,
            '未排班': -1,
            '未知班次': -2
        }
        
        for cat in result:
            result[cat].sort(key=lambda x: (
                -status_priority.get(x['status'], -3),
                self.get_shift_start_time(x['shift'])
            ))
        
        return result
    
    def get_shift_start_time(self, shift_code):
        if not shift_code or str(shift_code).strip() == '':
            return time(23, 59, 59)
            
        shift_code = str(shift_code).strip()
        
        for s in self.shift_times:
            if s in shift_code:
                return self.shift_times[s]['start']
        
        return time(23, 59, 59)

def download_from_jiananguo():
    try:
        jiananguo_email = st.secrets.get("JIANANGUO_EMAIL", "hanyong@foxmail.com")
        jiananguo_password = st.secrets.get("JIANANGUO_PASSWORD", "ah5fb6yahy62b8rt")
        
        options = {
            'webdav_hostname': 'https://dav.jianguoyun.com/dav/',
            'webdav_login': jiananguo_email,
            'webdav_password': jiananguo_password
        }
        
        client = Client(options)
        remote_file = '我的坚果云/排班.xlsx'
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            local_file = tmp_file.name
        
        client.download_sync(remote_path=remote_file, local_path=local_file)
        
        if os.path.exists(local_file) and os.path.getsize(local_file) > 0:
            return True, local_file, ""
        else:
            return False, None, "从坚果云下载文件失败"
            
    except Exception as e:
        return False, None, f"下载失败: {str(e)}"

def create_clickable_agent_card(person_info, viewer):
    """创建可点击的坐席卡片"""
    status_icon = viewer.status_icons.get(person_info['status'], '❓')
    
    # 状态颜色
    if person_info['status'] in ["正在路上", "已回家"]:
        status_color = "#BFBFBF"
        bg_color = "#F5F5F5"
        border_color = "#DDD"
    else:
        status_color = person_info['status_color']
        seat_type = person_info.get('actual_seat', person_info['seat'])
        bg_color = f"#{person_info['color']}" if seat_type in ['B席', 'C席'] else "#FFFFFF"
        border_color = "#333"
    
    # 添加点击效果和手型光标
    card_html = f"""
    <div style="background-color: {bg_color}; border: 1px solid {border_color}; border-radius: 4px; padding: 6px; margin: 2px; min-height: 60px; display: flex; flex-direction: column; justify-content: center; cursor: pointer; transition: all 0.2s ease;"
         onmouseover="this.style.transform='scale(1.02)'; this.style.boxShadow='0 2px 8px rgba(0,0,0,0.15)';"
         onmouseout="this.style.transform='scale(1)'; this.style.boxShadow='none';"
         onclick="handleCardClick('{person_info['id']}', '{person_info['name']}')">
        <div style="font-size: 14px; font-weight: bold; text-align: center; margin-bottom: 4px;">{person_info['name']}</div>
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div style="font-size: 12px; color: #666;">{person_info['workplace']}</div>
            <div style="font-size: 16px;">{status_icon}</div>
        </div>
        <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 2px;">
            <div style="font-size: 11px; font-weight: bold;">{person_info['shift']}</div>
            <div style="font-size: 11px; color: {status_color}; font-weight: bold;">{person_info['status']}</div>
        </div>
    </div>
    """
    
    return card_html

def create_month_schedule_calendar(month_schedule_data, current_date):
    """创建月度排班日历视图"""
    agent_info = month_schedule_data['agent_info']
    schedule = month_schedule_data['month_schedule']
    
    # 创建日历标题
    st.markdown(f"""
    <div style="background-color: #2E8B57; color: white; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
        <h2 style="margin: 0; text-align: center;">{agent_info['name']} - {current_date.year}年{current_date.month}月排班表</h2>
        <p style="margin: 5px 0 0 0; text-align: center;">工号: {agent_info['id']} | 职场: {agent_info['workplace']}</p>
    </div>
    """, unsafe_allow_html=True)
    
    # 创建日历
    cal = calendar.monthcalendar(current_date.year, current_date.month)
    
    # 星期标题
    weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    
    # 将排班数据转换为字典以便快速查找
    schedule_dict = {item['date']: item for item in schedule}
    
    # 创建日历表格
    calendar_html = """
    <div style="width: 100%; border-collapse: collapse;">
        <div style="display: flex; background-color: #f0f0f0; border: 1px solid #ddd;">
    """
    
    # 添加星期标题
    for day in weekdays:
        calendar_html += f"""
            <div style="flex: 1; padding: 10px; text-align: center; font-weight: bold; border-right: 1px solid #ddd;">
                {day}
            </div>
        """
    
    calendar_html += """
        </div>
    """
    
    # 添加日期和排班信息
    for week in cal:
        calendar_html += '<div style="display: flex; border: 1px solid #ddd; border-top: none;">'
        
        for i, day in enumerate(week):
            if day == 0:
                # 空日期
                calendar_html += '<div style="flex: 1; min-height: 80px; border-right: 1px solid #ddd; background-color: #f9f9f9;"></div>'
            else:
                current_day = date(current_date.year, current_date.month, day)
                day_schedule = schedule_dict.get(current_day, None)
                
                # 判断是否是今天
                today = datetime.now(TZ_UTC_8).date()
                is_today = current_day == today
                
                # 基础样式
                day_style = "flex: 1; min-height: 80px; padding: 5px; border-right: 1px solid #ddd;"
                
                if is_today:
                    day_style += "background-color: #fffacd; font-weight: bold;"
                elif current_day.weekday() >= 5:  # 周末
                    day_style += "background-color: #f0f8ff;"
                else:
                    day_style += "background-color: white;"
                
                calendar_html += f'<div style="{day_style}">'
                calendar_html += f'<div style="font-size: 14px; margin-bottom: 5px;">{day}</div>'
                
                if day_schedule:
                    shift_color = day_schedule['color']
                    seat_type = day_schedule['seat']
                    bg_color = f"#{shift_color}" if seat_type in ['B席', 'C席'] else "#FFFFFF"
                    
                    calendar_html += f"""
                    <div style="background-color: {bg_color}; padding: 3px; border-radius: 3px; font-size: 12px; text-align: center; border: 1px solid #ccc;">
                        <div style="font-weight: bold;">{day_schedule['shift']}</div>
                        <div style="font-size: 10px;">{day_schedule['seat']}</div>
                    </div>
                    """
                else:
                    calendar_html += '<div style="color: #999; font-size: 11px; text-align: center;">无排班</div>'
                
                calendar_html += '</div>'
        
        calendar_html += '</div>'
    
    calendar_html += '</div>'
    
    st.markdown(calendar_html, unsafe_allow_html=True)
    
    # 添加排班统计
    if schedule:
        st.subheader("排班统计")
        shift_counts = {}
        for item in schedule:
            shift = item['shift']
            shift_counts[shift] = shift_counts.get(shift, 0) + 1
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("总排班天数", len(schedule))
        with col2:
            st.metric("不同班次数", len(shift_counts))
        with col3:
            # 计算休息日
            rest_days = sum(1 for item in schedule if item['shift'] in ['休', '休息'])
            st.metric("休息日", rest_days)
        
        # 显示班次分布
        st.write("班次分布:")
        for shift, count in shift_counts.items():
            st.write(f"- {shift}: {count}天")

def update_current_time():
    weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
    now = datetime.now(TZ_UTC_8)
    weekday = weekdays[now.weekday()]
    return now.strftime(f"%Y年%m月%d日 {weekday} %H:%M:%S")

def auto_refresh_time(placeholder):
    while True:
        if not st.session_state.get('auto_refresh', True):
            t.sleep(1)
            continue
        placeholder.markdown(f"### 当前时间: {update_current_time()}")
        
        current_minute = datetime.now(TZ_UTC_8).minute
        if current_minute == 0 and not st.session_state.get('hour_refresh_done', False):
            st.session_state.hour_refresh_done = True
            st.session_state.refresh_counter += 1
            st.session_state.schedule_data = {}
            st.rerun()
        elif current_minute != 0:
            st.session_state.hour_refresh_done = False
            
        t.sleep(1)

def filter_data_by_workplace(df, workplace):
    """根据职场筛选数据"""
    if workplace == "全部":
        return df
    elif workplace in ["重庆", "北京"]:
        return df[df['workplace'] == workplace]
    else:
        return df

def filter_data_by_name(df, name_query):
    """根据姓名查询筛选数据"""
    if not name_query:
        return df
    return df[df['name'].str.contains(name_query, case=False, na=False)]

def show_agent_detail(viewer, agent_id, agent_name):
    """显示人员详情页面"""
    # 返回按钮
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("← 返回", use_container_width=True):
            st.session_state.show_detail = False
            st.session_state.detail_agent_id = None
            st.rerun()
    
    with col2:
        st.title(f"📅 {agent_name} 的月度排班")
    
    st.markdown("---")
    
    # 选择月份
    current_date = datetime.now(TZ_UTC_8)
    selected_month = st.selectbox(
        "选择月份",
        options=[
            (current_date.year, current_date.month),
            (current_date.year, current_date.month - 1) if current_date.month > 1 else (current_date.year - 1, 12),
            (current_date.year, current_date.month + 1) if current_date.month < 12 else (current_date.year + 1, 1)
        ],
        format_func=lambda x: f"{x[0]}年{x[1]}月",
        index=0,
        key="month_selector"
    )
    
    year, month = selected_month
    
    # 加载月度排班数据
    with st.spinner(f"正在加载 {year}年{month}月 的排班数据..."):
        month_schedule_data = viewer.load_agent_month_schedule(
            st.session_state.file_path, 
            agent_id, 
            year, 
            month
        )
    
    if month_schedule_data and month_schedule_data['month_schedule']:
        create_month_schedule_calendar(month_schedule_data, date(year, month, 1))
    else:
        st.warning(f"未找到 {agent_name} 在 {year}年{month}月 的排班数据")

def main():
    st.set_page_config(
        page_title="综合组在线坐席", 
        layout="wide",
        page_icon="📊"
    )
    
    # 简化session state管理
    if 'file_path' not in st.session_state:
        st.session_state.file_path = None
    if 'last_download' not in st.session_state:
        st.session_state.last_download = None
    if 'refresh_counter' not in st.session_state:
        st.session_state.refresh_counter = 0
    if 'workplace_filter' not in st.session_state:
        st.session_state.workplace_filter = "全部"
    if 'name_query' not in st.session_state:
        st.session_state.name_query = ""
    if 'show_detail' not in st.session_state:
        st.session_state.show_detail = False
    if 'detail_agent_id' not in st.session_state:
        st.session_state.detail_agent_id = None
    if 'detail_agent_name' not in st.session_state:
        st.session_state.detail_agent_name = None
    
    # 初始化查看器
    viewer = AgentViewer()
    
    # 首次运行或文件不存在时下载排班文件
    if st.session_state.file_path is None or not os.path.exists(st.session_state.file_path):
        with st.spinner("正在加载排班文件..."):
            download_success, file_path, download_message = download_from_jiananguo()
            if download_success:
                st.session_state.file_path = file_path
                st.session_state.last_download = datetime.now(TZ_UTC_8)
            else:
                st.error(f"加载失败: {download_message}")
                st.stop()
    
    # 如果显示详情页面，则显示详情
    if st.session_state.show_detail and st.session_state.detail_agent_id:
        show_agent_detail(viewer, st.session_state.detail_agent_id, st.session_state.detail_agent_name)
        return
    
    # 添加JavaScript处理卡片点击
    st.markdown("""
    <script>
    function handleCardClick(agentId, agentName) {
        // 发送数据到Streamlit
        window.parent.postMessage({
            type: 'streamlit:setComponentValue',
            value: {
                agent_id: agentId,
                agent_name: agentName
            }
        }, '*');
    }
    
    // 监听来自Streamlit的消息
    window.addEventListener('message', function(event) {
        if (event.data.type === 'streamlit:componentValue') {
            // 处理组件值变化
        }
    });
    </script>
    """, unsafe_allow_html=True)
    
    # 主界面 - 简化布局
    col_logo, col_title = st.columns([1, 4])
    
    with col_logo:
        logo_html = """
        <div style="display: flex; align-items: center; justify-content: center; padding: 5px;">
            <div style="text-align: center;">
                <h2 style="margin: 0; color: #2E8B57; font-weight: bold;">HealthLink</h2>
                <p style="margin: 0; color: #2E8B57; font-size: 12px;">远盟康健®</p>
            </div>
        </div>
        """
        st.markdown(logo_html, unsafe_allow_html=True)
    
    with col_title:
        st.title("综合组在线坐席")
    
    # 简化控制栏
    col_controls = st.columns([2, 1, 1, 1])
    
    with col_controls[0]:
        # 搜索和筛选区域
        col_search, col_filter = st.columns([2, 1])
        with col_search:
            name_query = st.text_input(
                "搜索姓名",
                placeholder="输入姓名关键字...",
                key=f"name_query_{st.session_state.refresh_counter}"
            )
            st.session_state.name_query = name_query
        
        with col_filter:
            workplace_filter = st.selectbox(
                "选择职场",
                ["全部", "重庆", "北京"],
                key=f"workplace_{st.session_state.refresh_counter}"
            )
            st.session_state.workplace_filter = workplace_filter
    
    with col_controls[1]:
        view_date = st.date_input(
            "选择日期", 
            datetime.now(TZ_UTC_8).date(),
            key=f"date_{st.session_state.refresh_counter}"
        )
    
    with col_controls[2]:
        hour_options = [f"{h:02d}:00" for h in range(24)]
        current_hour_str = f"{datetime.now(TZ_UTC_8).hour:02d}:00"
        
        default_idx = hour_options.index(current_hour_str) if current_hour_str in hour_options else 0
        
        selected_time_str = st.selectbox(
            "选择时间", 
            hour_options,
            index=default_idx,
            key=f"time_{st.session_state.refresh_counter}"
        )
        
        hour = int(selected_time_str.split(":")[0])
        view_time = time(hour, 0)
    
    with col_controls[3]:
        col_refresh1, col_refresh2 = st.columns(2)
        with col_refresh1:
            if st.button("🔄 刷新", use_container_width=True):
                st.session_state.refresh_counter += 1
                st.success("状态已刷新")
        
        with col_refresh2:
            if st.button("📥 重载", use_container_width=True):
                with st.spinner("重新加载中..."):
                    download_success, file_path, download_message = download_from_jiananguo()
                    if download_success:
                        st.session_state.file_path = file_path
                        st.session_state.last_download = datetime.now(TZ_UTC_8)
                        st.session_state.refresh_counter += 1
                        st.success("数据已更新")
                    else:
                        st.error(f"加载失败: {download_message}")
    
    st.markdown("---")
    
    # 显示当前查看时间
    weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
    weekday = weekdays[view_date.weekday()]
    
    # 简化数据加载逻辑
    current_hour = view_time.hour
    
    # 确定加载日期
    if current_hour < 8:
        # 8点之前显示前一天的排班
        load_date = view_date - timedelta(days=1)
        st.info(f"当前查看: {view_date.strftime('%Y年%m月%d日')} {weekday} {view_time.strftime('%H:%M')} (显示{load_date.strftime('%Y年%m月%d日')}的排班数据)")
    else:
        # 8点及之后显示当天的排班
        load_date = view_date
        st.info(f"当前查看: {view_date.strftime('%Y年%m月%d日')} {weekday} {view_time.strftime('%H:%M')}")
    
    # 使用日期字符串作为缓存键
    load_date_key = load_date.strftime('%Y-%m-%d')
    
    # 加载对应日期的数据
    if f"schedule_{load_date_key}" not in st.session_state:
        with st.spinner(f"正在加载{load_date.strftime('%Y年%m月%d日')}的坐席数据，请稍候..."):
            schedule_df = viewer.load_schedule_with_colors(
                st.session_state.file_path, 
                load_date
            )
            st.session_state[f"schedule_{load_date_key}"] = schedule_df
    else:
        schedule_df = st.session_state[f"schedule_{load_date_key}"]
    
    if schedule_df is None or schedule_df.empty:
        st.warning(f"未找到有效坐席数据")
        return
    
    # 应用职场筛选
    schedule_df = filter_data_by_workplace(schedule_df, st.session_state.workplace_filter)
    
    # 应用姓名查询
    schedule_df = filter_data_by_name(schedule_df, st.session_state.name_query)
    
    if schedule_df.empty:
        st.warning(f"未找到符合条件的坐席数据")
        return
    
    # 按A/B/C席分类显示坐席
    categorized_data = viewer.categorize_by_seat(schedule_df, view_time)
    
    # 看板式布局 - 三列并排
    st.subheader(f"{view_date.strftime('%Y年%m月%d日')} {weekday} 坐席看板")
    
    # 创建三列
    col_a, col_b, col_c = st.columns(3)
    
    # A席看板
    with col_a:
        agents_a = categorized_data.get('A席', [])
        online_count_a = sum(1 for agent in agents_a if agent['status'] == '搬砖中')
        total_count_a = len(agents_a)
        
        # 席位标题
        st.markdown(f"""
        <div style="background-color: #FFFFFF; border: 2px solid #333; border-radius: 6px; padding: 8px; margin-bottom: 8px; text-align: center;">
            <h3 style="margin: 0; color: #333;">A席 ({online_count_a}/{total_count_a})</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # 坐席网格 - 每行显示2个坐席
        if agents_a:
            # 计算每行显示的坐席数量
            cols_per_row = 2
            for i in range(0, len(agents_a), cols_per_row):
                cols = st.columns(cols_per_row)
                for j in range(cols_per_row):
                    if i + j < len(agents_a):
                        with cols[j]:
                            card_html = create_clickable_agent_card(agents_a[i + j], viewer)
                            st.markdown(card_html, unsafe_allow_html=True)
                            
                            # 添加点击处理
                            if st.button(f"查看 {agents_a[i + j]['name']} 的排班", 
                                       key=f"detail_{agents_a[i + j]['id']}_{i + j}",
                                       use_container_width=True):
                                st.session_state.show_detail = True
                                st.session_state.detail_agent_id = agents_a[i + j]['id']
                                st.session_state.detail_agent_name = agents_a[i + j]['name']
                                st.rerun()
        else:
            st.info("暂无A席坐席")
    
    # B席看板
    with col_b:
        agents_b = categorized_data.get('B席', [])
        online_count_b = sum(1 for agent in agents_b if agent['status'] == '搬砖中')
        total_count_b = len(agents_b)
        
        # 席位标题
        st.markdown(f"""
        <div style="background-color: #EF949F; border: 2px solid #333; border-radius: 6px; padding: 8px; margin-bottom: 8px; text-align: center;">
            <h3 style="margin: 0; color: #333;">B席 ({online_count_b}/{total_count_b})</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # 坐席网格 - 每行显示2个坐席
        if agents_b:
            # 计算每行显示的坐席数量
            cols_per_row = 2
            for i in range(0, len(agents_b), cols_per_row):
                cols = st.columns(cols_per_row)
                for j in range(cols_per_row):
                    if i + j < len(agents_b):
                        with cols[j]:
                            card_html = create_clickable_agent_card(agents_b[i + j], viewer)
                            st.markdown(card_html, unsafe_allow_html=True)
                            
                            # 添加点击处理
                            if st.button(f"查看 {agents_b[i + j]['name']} 的排班", 
                                       key=f"detail_{agents_b[i + j]['id']}_{i + j}",
                                       use_container_width=True):
                                st.session_state.show_detail = True
                                st.session_state.detail_agent_id = agents_b[i + j]['id']
                                st.session_state.detail_agent_name = agents_b[i + j]['name']
                                st.rerun()
        else:
            st.info("暂无B席坐席")
    
    # C席看板
    with col_c:
        agents_c = categorized_data.get('C席', [])
        online_count_c = sum(1 for agent in agents_c if agent['status'] == '搬砖中')
        total_count_c = len(agents_c)
        
        # 席位标题
        st.markdown(f"""
        <div style="background-color: #FFC000; border: 2px solid #333; border-radius: 6px; padding: 8px; margin-bottom: 8px; text-align: center;">
            <h3 style="margin: 0; color: #333;">C席 ({online_count_c}/{total_count_c})</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # 坐席网格 - 每行显示2个坐席
        if agents_c:
            # 计算每行显示的坐席数量
            cols_per_row = 2
            for i in range(0, len(agents_c), cols_per_row):
                cols = st.columns(cols_per_row)
                for j in range(cols_per_row):
                    if i + j < len(agents_c):
                        with cols[j]:
                            card_html = create_clickable_agent_card(agents_c[i + j], viewer)
                            st.markdown(card_html, unsafe_allow_html=True)
                            
                            # 添加点击处理
                            if st.button(f"查看 {agents_c[i + j]['name']} 的排班", 
                                       key=f"detail_{agents_c[i + j]['id']}_{i + j}",
                                       use_container_width=True):
                                st.session_state.show_detail = True
                                st.session_state.detail_agent_id = agents_c[i + j]['id']
                                st.session_state.detail_agent_name = agents_c[i + j]['name']
                                st.rerun()
        else:
            st.info("暂无C席坐席")

if __name__ == "__main__":
    main()
